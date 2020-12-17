import requests
import os
from os.path import expanduser
import logging
import time
import json
from openpyxl import Workbook
from datetime import datetime

logFile = expanduser("~") + "/Documents/Python_Projects_GIT/Parsable_Scripts/" + "LogFile.log"
logging.basicConfig(level = logging.DEBUG, filename = logFile, format='%(asctime)s - %(levelname)s - %(message)s')

class Parsable():
	def __init__(self):
		"""Initialize all useful information"""

		# Environment Team IDs
		self.production_team_id = "f0b1ad0a-e589-413e-bf9d-3fdaf7590df8"
		self.sandbox_team_id = "7bd5118a-9e55-44e7-992d-ba30854e31ef" 

		# Permanent Authorization Token (Production)
		self.authToken = 
		
		# Temperary Authorization Token (Sandbox - Update 30 days)
		# self.authToken = 
		
		# API base URL
		self.baseURL = "https://api.parsable.net/api/"
		# baseURL = "http://localhost:8080/api/"

		# API Headers
		self.headers = {
		    "accept": "application/json",
		    "cache-control": "no-cache",
		    "content-type": "application/json",
		    "Authorization": self.authToken
		}

		# List of templates. [Lip Print Template]
		self.template_id_list = ["9babcbd6-b34f-433c-a78f-ff0f976d354c"]

	def query_templates(self):
		"""
		Querys a list of template objects that contain a template's metadata.

		Parameters:
			NONE

		Returns:
			Template Object List: A list of template objects that contain a template's metadata.
		"""

		# Build an API call which querys a list of all templates.
		url = self.baseURL + "job_templates#query"
		payload = {
		    "method": "query",
    		"arguments": {
	        	"selectOpts": {
		            "includeTeam": True,
		            "includeRootHeaders": False,
		            "includeSteps": False,
		            "includeDocuments": False,
		            "includeLastPublished": False,
		            "includeLastAuthor": False,
		            "includeStats": False,
		            "includeTags": False,
		            "includeDrafts": False,
		            "includeLastModified": False,
		            "includeAttributes": False,
		            "includeRefMap": False,
		            "includeOriginalAuthor": False
		        },
	        	"whereOpts": {
	            	"teamId": self.production_team_id,
	            	"isArchived": False
	        	}
    		}
		}

		# Get template data Request
		template_list_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)

		# Make sure request was successful
		if template_list_response.status_code == 200:
			# Get JSON response object
			jsonRespose = template_list_response.json()

			# Make sure JSON object is good
			if not ("err" in jsonRespose['result']) and not ("exception" in jsonRespose):
				# Return a template list from json response
				return jsonRespose['result']['success']["templates"]
				
			# TODO: Fix to match if statement			
			# else:
			# 	logging.error(str(jsonRespose["result"]["err"]["errorCode"]) + " - JSON Object Error!")
		else:
			logging.error(str(template_list_response.status_code) + " - GetData POST Request Unsuccessful!")

	def query_jobs(self, templateIds):
		"""
		Querys a list of job objects based on included templates that contain a job's metadata.

		Parameters:
			templateIds (list): List of template IDs.

		Returns:
			Job Object List: A list of job objects that contain a job's metadata.
		"""

		# Build an API call which querys a list of jobs based on the template ID value.
		url = self.baseURL + "jobs#query"
		payload = {
		    "method": "query",
		    "arguments": {
		        "selectOpts": {
		            "includeTeam": False,
		            "includeTemplate": False,
		            "includeRootHeaders": False,
		            "includeSteps": False,
		            "includeDocuments": False,
		            "includeUsers": False,
		            "includeStats": False,
		            "includeActivity": False,
		            "includeTemplates": False,
		            "includeCreator": False,
		            "includeRoles": False,
		            "includePermissions": False,
		            "includeExecSnippets": False,
		            "includeMessages": False,
		            "includeIssues": False,
		            "includeDeviationCounts": False,
		            "includeDeviations": False,
		            "includeRefMap": False,
		            "includePlannedDataSheetIds": False,
		            "includeSnapshottedDataSheetValues": False,
		            "includeAttributes": False
		        },
		        "whereOpts": {
		            "teamId": self.production_team_id,
		            "isComplete": True,
		            "sourceTemplateIds": templateIds,
		            "allSourceTemplateIds": True
		        }
		    }
		}

		# Get job data Request
		job_list_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)

		# Make sure request was successful
		if job_list_response.status_code == 200:
			# Get JSON response object
			jsonRespose = job_list_response.json()

			# Make sure JSON object is good
			if not ("err" in jsonRespose['result']) and not ("exception" in jsonRespose):
				# Return a job list from json response
				return jsonRespose['result']['success']["jobs"]
				
			# TODO: Fix to match if statement			
			# else:
			# 	logging.error(str(jsonRespose["result"]["err"]["errorCode"]) + " - JSON Object Error!")
		else:
			logging.error(str(job_list_response.status_code) + " - GetData POST Request Unsuccessful!")

	def query_jobs_before_date(self, unixTimeStamp):
		"""
		Querys a list of job objects based on creation date that contain a job's metadata.

		Parameters:
			unixTimeStamp (int): Unix time stamp of before date

		Returns:
			Job Object List: A list of job objects that contain a job's metadata.
		"""

		# Build an API call which querys a list of jobs based on the template ID value.
		url = self.baseURL + "jobs#query"
		payload = {
		    "method": "query",
		    "arguments": {
		        "selectOpts": {
		            "includeTeam": False,
		            "includeTemplate": False,
		            "includeRootHeaders": False,
		            "includeSteps": False,
		            "includeDocuments": False,
		            "includeUsers": False,
		            "includeStats": False,
		            "includeActivity": False,
		            "includeTemplates": False,
		            "includeCreator": False,
		            "includeRoles": False,
		            "includePermissions": False,
		            "includeExecSnippets": False,
		            "includeMessages": False,
		            "includeIssues": False,
		            "includeDeviationCounts": False,
		            "includeDeviations": False,
		            "includeRefMap": False,
		            "includePlannedDataSheetIds": False,
		            "includeSnapshottedDataSheetValues": False,
		            "includeAttributes": False
		        },
		        "whereOpts": {
		            "teamId": self.production_team_id,
		            "isComplete": False,
		            "isStarted": True,
		            "startedBeforeTime": unixTimeStamp
		        }
		    }
		}

		# Get job data Request
		job_list_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)

		# Make sure request was successful
		if job_list_response.status_code == 200:
			# Get JSON response object
			jsonRespose = job_list_response.json()

			# Make sure JSON object is good
			if not ("err" in jsonRespose['result']) and not ("exception" in jsonRespose):
				# Return a job list from json response
				return jsonRespose['result']['success']["jobs"]
				
			# TODO: Fix to match if statement			
			# else:
			# 	logging.error(str(jsonRespose["result"]["err"]["errorCode"]) + " - JSON Object Error!")
		else:
			logging.error(str(job_list_response.status_code) + " - GetData POST Request Unsuccessful!")

	def get_job_data(self, jobObject):
		"""
		Get the data collected from this job.

		Parameters:
			jobObject (Object): Job metadata object used to make a request to get actual job data.

		Returns:
			Job Data (Object): Job object that contains all collected data inside of job.
		"""

		logging.info(jobObject["lookupId"] + " - Started")
		beforeJob = time.time()
		# Check to see if job is already downloaded
		if not os.path.exists(expanduser("~") + "/Desktop/PythonImages/" + jobObject["lookupId"]):
			# Build the first API call which gets the job data based on the JobUUID value.
			url = self.baseURL + "jobs#getData"
			payload = {
			    "method": "getData",
			    "arguments": {
			        "jobId": jobObject["id"],
			        "seqId": 1,
			        "options": {
			        	"canHandlePendingDocuments": True
			        }
			    }
			}

			# Get job data Request
			job_data_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)

			# Make sure request was successful
			if job_data_response.status_code == 200:
				# Get JSON response object
				jsonRespose = job_data_response.json()

				# Make sure JSON object is good
				if not ("err" in jsonRespose['result']) and not ("exception" in jsonRespose):

					# Return job data from json response
					return jsonRespose['result']['success']

				# TODO: Fix to match if statement
				# else:
				# 	logging.error(str(jsonRespose["result"]["err"]["errorCode"]) + " - JSON Object Error!")
			else:
				logging.error(str(job_data_response.status_code) + " - GetData POST Request Unsuccessful!")
		afterJob = time.time()
		logging.info(jobObject["lookupId"] + " - Already Downloaded")

	def get_all_document_ids(self, jobData, jobLookupId):
		"""
		Get all document IDs inside of this job

		Parameters:
			jobLookupId (string): Look up ID of the specific job. I.e. Job-xxxxx
			jobData (Object): Job data object that includes all collected data from job.

		Returns:
			Document List: A list of tuples that contain (jobLookupId, documentId, fieldId)
		"""

		#List of Tuples that contain (jobLookupId, documentId, fieldId)
		document_list = []

		# GRAB ALL IMAGE IDs IN JOB
		# Loop through all templates in the job
		for template in jobData["snippets"]:
			# Make sure it is a regular step
			# TODO: Add else if statements if needd for StepGroup
			if "stepExecData" in template:
				# Loop through all regular steps in the template, excluding 
				for step in template["stepExecData"]["fieldExecutionData"]:
					# CHECK to see if it is an image input
					if "documents" in step:
						# Grab the first image ID in the image list
						imageId = step["documents"][0]["id"]
						# Tuple that contains (jobLookupId, documentId, fieldId)
						documentTuple = (jobLookupId, imageId, step["fieldId"])
						# Append document Tuple to the document list
						document_list.append(documentTuple)
						# Download Image
						# Call can be moved outside of this function. Only here to speed things up.
						self.download_document(documentTuple)
					# CHECK to see if it is an signature input
					elif "document" in step:
						# Grab Signature ID
						signatureId = step["document"]["id"]
						# Tuple that contains (jobLookupId, documentId, fieldId)
						documentTuple = (jobLookupId, signatureId, step["fieldId"])
						# Append document Tuple to the document list
						document_list.append(documentTuple)
						# Download Signature
						# Call can be moved outside of this function. Only here to speed things up.
						self.download_document(documentTuple)
					else:
						logging.info("Not an image.")
		return document_list

	def download_document(self, documentInfo):
		"""
		Download document based on the ID.

		Parameters:
			documentInfo (Tuple): Tuple that contains (jobLookupId, documentId, fieldId).
		"""
		# If folder doen't exist create it then download image else download image to existng folder
		if not os.path.exists(expanduser("~") + "/Desktop/PythonImages/" + documentInfo[0]):
			# Create folder to hold all images related to this job
			os.mkdir(expanduser("~") + "/Desktop/PythonImages/" + documentInfo[0])

		# Image request URL
		documentUrl = self.baseURL + "documents/" + documentInfo[1]

		# GET image from request
		document_response = requests.request('GET', url=documentUrl, data=None, json=None, headers=self.headers)

		# Save image from json response
		# Download image from requests
		if document_response.status_code == 200:
			# Save image with name: JobID_FieldID.jpeg in JobID folder
			destinationName = expanduser("~") + "/Desktop/PythonImages/" + documentInfo[0] + "/" + documentInfo[0] + "_" + documentInfo[2] + ".jpeg"
			
			# w - write and b - binary (images)
			with open(destinationName, "wb") as x:
				x.write(document_response.content)
		else:
			logging.error(response.status_code + " - Image GET Request Unsuccessful!")

	def template_metadata_to_excel(self, template_list):
		filename = "Corteva_Production_Templates.xlsx"
		column_names = ["Template Version",
						"Last Published",
						"Template Title",
						"Template Description",
						"Region",
						"Location",
						"Status",
						"Department",
						"Process"]

		workbook = Workbook()
		sheet = workbook.active

		# Create title row
		for i in range(0, len(column_names)):
			# Row & Col start @ 1
			sheet.cell(row = 1, column = i+1, value = column_names[i])

		# Loop through all templates in template list.
		for i in range(0, len(template_list)):
			template = template_list[i]
			for j in range(0, len(column_names)):
				if column_names[j] == "Template Version":
					# Insert Template Version
					# Start on row 2 as row 1 is headers.
					sheet.cell(row = i+2, column = j+1, value = template["publicVersion"])
				elif column_names[j] == "Last Published":
					# Insert Last Published
					sheet.cell(row = i+2, column = j+1, value = datetime.fromtimestamp(template["publishedAt"]))
				elif column_names[j] == "Template Title":
					# Insert Template Title
					sheet.cell(row = i+2, column = j+1, value = template["title"])
				elif column_names[j] == "Template Description":
					# Insert Template Description
					sheet.cell(row = i+2, column = j+1, value = template["descrip"])
				elif column_names[j] == "Region":
					if "1ca6eba0-7aba-4b68-bb29-c99676450b84" in template["metadataById"].keys():
						# Create a string with all Regions
						region_string = ""
						for region in template["metadataById"]["1ca6eba0-7aba-4b68-bb29-c99676450b84"]["value"]["selectedListValues"]["values"]:
							region_string += region + ", "
						# Insert Region
						sheet.cell(row = i+2, column = j+1, value = region_string)
					else:
						sheet.cell(row = i+2, column = j+1, value = "N/A")
				elif column_names[j] == "Location":
					if "02bf6962-5658-442d-b5bf-51b48f5e7062" in template["metadataById"].keys():
						# Create a string with all Locations
						location_string = ""
						for location in template["metadataById"]["02bf6962-5658-442d-b5bf-51b48f5e7062"]["value"]["selectedListValues"]["values"]:
							location_string += location + ", "
						# Insert Location
						sheet.cell(row = i+2, column = j+1, value = location_string)
					else:
						sheet.cell(row = i+2, column = j+1, value = "N/A")
				elif column_names[j] == "Status":
					# Insert Status
					sheet.cell(row = i+2, column = j+1, value = "Production")
				elif column_names[j] == "Department":
					if "80ac5e3a-3ba2-41f2-8714-2e7e53d2fd12" in template["metadataById"].keys():
						# Insert Department
						sheet.cell(row = i+2, column = j+1, value = template["metadataById"]["80ac5e3a-3ba2-41f2-8714-2e7e53d2fd12"]["value"]["listValue"])
					else:
						sheet.cell(row = i+2, column = j+1, value = "N/A")
				elif column_names[j] == "Process":
					if "5d28ab79-edd9-491b-be10-2d68129eee79" in template["metadataById"].keys():
						# Insert Process
						sheet.cell(row = i+2, column = j+1, value = template["metadataById"]["5d28ab79-edd9-491b-be10-2d68129eee79"]["value"]["listValue"])
					else:
						sheet.cell(row = i+2, column = j+1, value = "N/A")
		
		workbook.save(filename = filename)

	def complete_job(self, jobId):
		"""
		Complete a job based on JobID.

		Parameters:
			jobId (String): Job ID String used to close a specific job.

		Returns:
			Status Code (String): JSON response status code
		"""

		# Build the first API call which completes a job based on jobID.
		url = self.baseURL + "jobs#complete"

		payload = {
			"method": "complete",
			"arguments": {
				"jobId": jobId,
				"reason": "Test"
			}
		}

		# Get job data Request
		job_completion_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)

		# Make sure request was successful
		if job_completion_response.status_code == 200:
			return str(job_completion_response.status_code)
		else:
			logging.error(str(job_completion_response.status_code) + " - GetData POST Request Unsuccessful!")
 
try:
	before_download = time.time()

	parsable = Parsable()
	
	templates = parsable.query_templates()
	parsable.template_metadata_to_excel(templates)

	# with open('parsable_jobs.json', 'w') as json_file:
	# 	json.dump(parsable.query_jobs_before_date(1593583200), json_file)

	"""
	Template Version -> ["publicVersion"]
	Last Published -> ["publishedAt"] "Epoch & Unix Timestamp"
	Template Title -> ["title"]
	Template Description -> ["descrip"]
	Region -> "1ca6eba0-7aba-4b68-bb29-c99676450b84"
	Location -> ""02bf6962-5658-442d-b5bf-51b48f5e7062"
	Status = P(Production)
	Department -> "80ac5e3a-3ba2-41f2-8714-2e7e53d2fd12"
	Process -> "5d28ab79-edd9-491b-be10-2d68129eee79"
	Document # -> [""] --> N/A
	Document Owner -> [""] --> N/A
	"""
	after_download = time.time()
	logging.info("Execution Finished: " + str(after_download - before_download) + " seconds")
except Exception as e:
	logging.exception("Exception occurred")








