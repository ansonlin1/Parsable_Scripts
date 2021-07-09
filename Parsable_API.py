import requests
import os
from os.path import expanduser
import logging
import time
from openpyxl import Workbook
from datetime import datetime

class Parsable():
	def __init__(self):
		"""Initialize all useful information"""

		self.production_team_id = "f0b1ad0a-e589-413e-bf9d-3fdaf7590df8"
		# self.sandbox_team_id = "7bd5118a-9e55-44e7-992d-ba30854e31ef" 

		#TODO: MUST ADD TOKEN
		self.authToken = 
		
		# List of templates. ["NA SAF Near Miss"]
		self.template_id_list = ["731a87e6-9b87-48d2-bde3-11fbb5e0595c"]

		# API base URL
		self.baseURL = "https://api.parsable.net/api/"
		##baseURL = "http://localhost:8080/api/"

		# API Headers
		self.headers = {
		    "accept": "application/json",
		    "cache-control": "no-cache",
		    "content-type": "application/json",
		    "Authorization": self.authToken
		}

		self.photo_dest = ""

	def query_jobs(self, templateIds):
		"""
		Querys a list of job objects within the last 24 hours that contain a specific list of templates.

		Parameters:
			templateIds (list): List of template IDs.

		Returns:
			Job Object List: A list of job objects that contain a jobs metadata.
		"""

		# Build an API call which querys a list of jobs based on the template ID value.
		url = self.baseURL + "jobs#query"
		payload = {
		    "method": "query",
		    "arguments": {
		        "selectOpts": {
		            "includeTeam": False,
		            "includeTemplate": False,
		            "includeRootHeaders": False,
		            "includeSteps": False,
		            "includeDocuments": False,
		            "includeUsers": False,
		            "includeStats": False,
		            "includeActivity": False,
		            "includeTemplates": False,
		            "includeCreator": False,
		            "includeRoles": False,
		            "includePermissions": False,
		            "includeExecSnippets": False,
		            "includeMessages": False,
		            "includeIssues": False,
		            "includeDeviationCounts": False,
		            "includeDeviations": False,
		            "includeRefMap": False,
		            "includePlannedDataSheetIds": False,
		            "includeSnapshottedDataSheetValues": False,
		            "includeAttributes": False
		        },
		        "whereOpts": {
		            "teamId": self.production_team_id,
		            "isComplete": True,
		            "sourceTemplateIds": templateIds,
					# Epoch Time: Current Time - 24 hour
                    "completedSinceTime": (int(time.mktime(time.localtime())) - 86400)
		        }
		    }
		}

		# Get job data Request
		job_list_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)
		# Make sure request was successful
		if job_list_response.status_code == 200:
			# Get JSON response object
			jsonRespose = job_list_response.json()
 			# Make sure JSON object is good
			if not ("err" in jsonRespose['result']) and not ("exception" in jsonRespose):
				# Return a job list from json response
				return jsonRespose['result']['success']["jobs"]
				
			# TODO: Fix to match if statement			
			# else:
			# 	logging.error(str(jsonRespose["result"]["err"]["errorCode"]) + " - JSON Object Error!")
		else:
			logging.error(str(job_list_response.status_code) + " - GetData POST Request Unsuccessful!")

	def query_jobs_last_day(self):
		"""
		Querys a list of all job objects within the last 24 hours.

		Returns:
			Job Object List: A list of job objects that contain a jobs metadata.
		"""

		# Build an API call which querys a list of all job objects within the last 24 hours.
		url = self.baseURL + "jobs#query"
		payload = {
		    "method": "query",
		    "arguments": {
		        "selectOpts": {
		            "includeTeam": False,
		            "includeTemplate": False,
		            "includeRootHeaders": False,
		            "includeSteps": False,
		            "includeDocuments": False,
		            "includeUsers": False,
		            "includeStats": False,
		            "includeActivity": False,
		            "includeTemplates": False,
		            "includeCreator": False,
		            "includeRoles": False,
		            "includePermissions": False,
		            "includeExecSnippets": False,
		            "includeMessages": False,
		            "includeIssues": False,
		            "includeDeviationCounts": False,
		            "includeDeviations": False,
		            "includeRefMap": False,
		            "includePlannedDataSheetIds": False,
		            "includeSnapshottedDataSheetValues": False,
		            "includeAttributes": False
		        },
		        "whereOpts": {
		            "teamId": self.production_team_id,
		            "isComplete": True,
                    # Epoch Time: Current Time - 24 hour
                    "completedSinceTime": (int(time.mktime(time.localtime())) - 86400),
		        }
		    }
		}

		# Get job data Request
		job_list_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)
		# Make sure request was successful
		if job_list_response.status_code == 200:
			# Get JSON response object
			jsonRespose = job_list_response.json()
 			# Make sure JSON object is good
			if not ("err" in jsonRespose['result']) and not ("exception" in jsonRespose):
				# Return a job list from json response
				return jsonRespose['result']['success']["jobs"]
				
			# TODO: Fix to match if statement			
			# else:
			# 	logging.error(str(jsonRespose["result"]["err"]["errorCode"]) + " - JSON Object Error!")
		else:
			logging.error(str(job_list_response.status_code) + " - GetData POST Request Unsuccessful!")

	def query_jobs_within_timeframe(self, since_day, before_day):
		"""
		Querys a list of all job within a specific timeframe.

		Parameters:
			since_day (int): Start Date
			before_day (int): End Date

		Returns:
			Job Object List: A list of job objects that contain a jobs metadata.
		"""

		# Build an API call which querys a list of all job within a specific timeframe.
		url = self.baseURL + "jobs#query"
		payload = {
			"method": "query",
			"arguments": {
				"selectOpts": {
					"includeTeam": False,
					"includeTemplate": False,
					"includeRootHeaders": False,
					"includeSteps": False,
					"includeDocuments": False,
					"includeUsers": False,
					"includeStats": False,
					"includeActivity": False,
					"includeTemplates": False,
					"includeCreator": False,
					"includeRoles": False,
					"includePermissions": False,
					"includeExecSnippets": False,
					"includeMessages": False,
					"includeIssues": False,
					"includeDeviationCounts": False,
					"includeDeviations": False,
					"includeRefMap": False,
					"includePlannedDataSheetIds": False,
					"includeSnapshottedDataSheetValues": False,
					"includeAttributes": False
				},
				"whereOpts": {
					"teamId": self.production_team_id,
					"isComplete": True,
					# Epoch Time
					"completedSinceTime": since_day,
					"completedBeforeTime": before_day
				}
			}
		}

		# Get job data Request
		job_list_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)
		# Make sure request was successful
		if job_list_response.status_code == 200:
			# Get JSON response object
			jsonRespose = job_list_response.json()
			# Make sure JSON object is good
			if not ("err" in jsonRespose['result']) and not ("exception" in jsonRespose):
				# Return a job list from json response
				return jsonRespose['result']['success']["jobs"]
				
			# TODO: Fix to match if statement			
			# else:
			# 	logging.error(str(jsonRespose["result"]["err"]["errorCode"]) + " - JSON Object Error!")
		else:
			logging.error(str(job_list_response.status_code) + " - GetData POST Request Unsuccessful!")

	def get_job_data(self, jobObject):
		"""
		Get the data collected from this job.

		Parameters:
			jobObject (Object): Job metadata object used to make a request to get actual job data.

		Returns:
			Job Data (Object): Job object that contains all collected data inside of job.
		"""

		logging.info(jobObject["lookupId"] + " - Started")
		# Check to see if job is already downloaded
		if os.path.exists(os.path.join(self.photo_dest, jobObject["lookupId"])):
			logging.info(jobObject["lookupId"] + " - Already Downloaded")
		else:
			# Create folder to hold all images related to this job
			os.mkdir(os.path.join(self.photo_dest, jobObject["lookupId"]))

			# Build the first API call which gets the job data based on the JobUUID value.
			url = self.baseURL + "jobs#getData"
			payload = {
			    "method": "getData",
			    "arguments": {
			        "jobId": jobObject["id"],
			        "seqId": 1,
			        "options": {
			        	"canHandlePendingDocuments": True
			        }
			    }
			}

			# Get job data Request
			job_data_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)

			# Make sure request was successful
			if job_data_response.status_code == 200:
				# Get JSON response object
				jsonRespose = job_data_response.json()

				# Make sure JSON object is good
				if not ("err" in jsonRespose['result']) and not ("exception" in jsonRespose):

					# Return job data from json response
					return jsonRespose['result']['success']

				# TODO: Fix to match if statement
				# else:
				# 	logging.error(str(jsonRespose["result"]["err"]["errorCode"]) + " - JSON Object Error!")
			else:
				logging.error(str(job_data_response.status_code) + " - GetData POST Request Unsuccessful!")

	def get_all_document_ids(self, jobData, jobLookupId):
		"""
		Get all document IDs inside of this job

		Parameters:
			jobLookupId (string): Look up ID of the specific job. I.e. Job-xxxxx
			jobData (Object): Job data object that includes all collected data from job.
		"""

		# GRAB ALL IMAGE IDs IN JOB
		# Loop through all steps in the job
		for job_step in jobData["snippets"]:
			# Make sure it is a regular step
			# TODO: Add else if statements if need for StepGroup
			if "stepExecData" in job_step:
				# Loop through step content (input)
				for job_input in job_step["stepExecData"]["fieldExecutionData"]:
					# CHECK to see if it is an image input
					if "documents" in job_input:
						for job_image in job_input["documents"]:
							# Grab the image ID in the image list
							imageId = job_image["id"]
							# Tuple that contains (jobLookupId, documentId)
							documentTuple = (jobLookupId, imageId)
							# Download Image
							# Call can be moved outside of this function. Only here to speed things up.
							self.download_document(documentTuple)
					# CHECK to see if it is an signature input
					elif "document" in job_input:
						# Grab Signature ID
						signatureId = job_input["document"]["id"]
						# Tuple that contains (jobLookupId, documentId)
						documentTuple = (jobLookupId, signatureId)
						# Download Signature
						# Call can be moved outside of this function. Only here to speed things up.
						self.download_document(documentTuple)
					else:
						logging.info("Not an image.")

	def download_document(self, documentInfo):
		"""
		Download document based on the ID.

		Parameters:
			documentInfo (Tuple): Tuple that contains (jobLookupId, documentId).
		"""

		# Image request URL
		documentUrl = self.baseURL + "documents/" + documentInfo[1]

		# GET image from request
		document_response = requests.request('GET', url=documentUrl, data=None, json=None, headers=self.headers)

		# Save image from json response
		# Download image from requests
		if document_response.status_code == 200:
			# Save image with name: JobID_FieldID.jpeg in JobID folder
			destinationName = os.path.join(self.photo_dest, documentInfo[0], (documentInfo[0] + "_" + documentInfo[1] + ".jpeg"))
			
			# w - write and b - binary (images)
			with open(destinationName, "wb") as x:
				x.write(document_response.content)
		else:
			logging.error(document_response.status_code + " - Image GET Request Unsuccessful!")

	def template_metadata_to_excel(self, template_list):
		filename = "Corteva_Production_Templates.xlsx"
		column_names = ["Template Version",
						"Last Published",
						"Template Title",
						"Template Description",
						"Region",
						"Location",
						"Status",
						"Department",
						"Process"]

		workbook = Workbook()
		sheet = workbook.active

		# Create title row
		for i in range(0, len(column_names)):
			# Row & Col start @ 1
			sheet.cell(row = 1, column = i+1, value = column_names[i])

		# Loop through all templates in template list.
		for i in range(0, len(template_list)):
			template = template_list[i]
			for j in range(0, len(column_names)):
				if column_names[j] == "Template Version":
					# Insert Template Version
					# Start on row 2 as row 1 is headers.
					sheet.cell(row = i+2, column = j+1, value = template["publicVersion"])
				elif column_names[j] == "Last Published":
					# Insert Last Published
					sheet.cell(row = i+2, column = j+1, value = datetime.fromtimestamp(template["publishedAt"]))
				elif column_names[j] == "Template Title":
					# Insert Template Title
					sheet.cell(row = i+2, column = j+1, value = template["title"])
				elif column_names[j] == "Template Description":
					# Insert Template Description
					sheet.cell(row = i+2, column = j+1, value = template["descrip"])
				elif column_names[j] == "Region":
					if "1ca6eba0-7aba-4b68-bb29-c99676450b84" in template["metadataById"].keys():
						# Create a string with all Regions
						region_string = ""
						for region in template["metadataById"]["1ca6eba0-7aba-4b68-bb29-c99676450b84"]["value"]["selectedListValues"]["values"]:
							region_string += region + ", "
						# Insert Region
						sheet.cell(row = i+2, column = j+1, value = region_string)
					else:
						sheet.cell(row = i+2, column = j+1, value = "N/A")
				elif column_names[j] == "Location":
					if "02bf6962-5658-442d-b5bf-51b48f5e7062" in template["metadataById"].keys():
						# Create a string with all Locations
						location_string = ""
						for location in template["metadataById"]["02bf6962-5658-442d-b5bf-51b48f5e7062"]["value"]["selectedListValues"]["values"]:
							location_string += location + ", "
						# Insert Location
						sheet.cell(row = i+2, column = j+1, value = location_string)
					else:
						sheet.cell(row = i+2, column = j+1, value = "N/A")
				elif column_names[j] == "Status":
					# Insert Status
					sheet.cell(row = i+2, column = j+1, value = "Production")
				elif column_names[j] == "Department":
					if "80ac5e3a-3ba2-41f2-8714-2e7e53d2fd12" in template["metadataById"].keys():
						# Insert Department
						sheet.cell(row = i+2, column = j+1, value = template["metadataById"]["80ac5e3a-3ba2-41f2-8714-2e7e53d2fd12"]["value"]["listValue"])
					else:
						sheet.cell(row = i+2, column = j+1, value = "N/A")
				elif column_names[j] == "Process":
					if "5d28ab79-edd9-491b-be10-2d68129eee79" in template["metadataById"].keys():
						# Insert Process
						sheet.cell(row = i+2, column = j+1, value = template["metadataById"]["5d28ab79-edd9-491b-be10-2d68129eee79"]["value"]["listValue"])
					else:
						sheet.cell(row = i+2, column = j+1, value = "N/A")
		
		workbook.save(filename = filename)
	#TODO: Not Complete
	def complete_job(self, jobId):
		"""
		Complete a job based on JobID.

		Parameters:
			jobId (String): Job ID String used to close a specific job.

		Returns:
			Status Code (String): JSON response status code
		"""

		# Build the first API call which completes a job based on jobID.
		url = self.baseURL + "jobs#complete"

		payload = {
			"method": "complete",
			"arguments": {
				"jobId": jobId,
				"reason": "Test"
			}
		}

		# Get job data Request
		job_completion_response = requests.request('POST', url=url, data=None, json=payload, headers=self.headers)

		# Make sure request was successful
		if job_completion_response.status_code == 200:
			return str(job_completion_response.status_code)
		else:
			logging.error(str(job_completion_response.status_code) + " - GetData POST Request Unsuccessful!")